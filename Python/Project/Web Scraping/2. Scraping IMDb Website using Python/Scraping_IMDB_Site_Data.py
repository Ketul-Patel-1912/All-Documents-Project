from bs4 import BeautifulSoup
import openpyxl
import requests

excel = openpyxl.Workbook()
print(excel.sheetnames) ## checking how many sheets allocated 

sheet = excel.active
sheet.title='Top Rated Movies'
print(excel.sheetnames)
sheet.append(['Movie_Rank','Movie_Name','Year_Of_Ralease','IMDB_Rating'])

try:
    ## handle code crash due to invalid link, prefer try block
    source  =  requests.get('https://www.imdb.com/chart/top/?ref_=nv_mv_250') ## get html information using request module
    source.raise_for_status() ## if source link not valid , received error
    
    soup = BeautifulSoup(source.text,'html.parser')
    movies = soup.find('tbody',class_='lister-list').find_all('tr')
    
    for movie in movies:
        
        name = movie.find('td',class_='titleColumn').a.text
        rank = movie.find('td',class_='titleColumn').get_text(strip=True).split('.')[0]
        year = movie.find('td',class_='titleColumn').span.text.strip("()")
        rating = movie.find('td',class_='ratingColumn imdbRating').strong.text
        sheet.append([rank,name,year,rating])
        
        
except Exception as e:
    print(e)
    
excel.save('IMDB Movie Ratings.xlsx')


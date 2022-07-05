from bs4 import BeautifulSoup
import requests
import openpyxl

exl = openpyxl.Workbook()
print(exl.sheetnames)
sheet = exl.active

sheet.title="Top Rated TV Show"
print(exl.sheetnames)

sheet.append(['TV_Show_Rank','TV_Show_Name','Year_Of_Ralease','IMDB_Rating'])

try:
    source = requests.get('https://www.imdb.com/chart/toptv/?ref_=nv_tvv_250')
    source.raise_for_status()
    
    soup = BeautifulSoup(source.text,'html.parser')
    movies = soup.find('tbody',class_='lister-list').find_all('tr')
    
    for movie in movies:
        name = movie.find('td',class_='titleColumn').a.text
        rank = movie.find('td',class_='titleColumn').get_text(strip=True).split(".")[0]
        year = movie.find('td',class_='titleColumn').span.text.strip('()')
        rating = movie.find('td',class_='ratingColumn imdbRating').strong.text
        
        sheet.append([rank,name,year,rating])
        
        
except Exception as e:
    print(e)

exl.save('Top Rated TV-Show.xlsx')